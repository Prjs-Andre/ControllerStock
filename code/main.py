import pandas as pd
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk

CAMINHO = r'planilha_estoque.xlsx'
DATA = pd.read_excel(CAMINHO, sheet_name="Estoque")


#
# Interface Gráfica
#
class ManagerInterface:
    def __init__(self):
        self.root = Tk()

        # Configurações Iniciais da Janela
        self.root.title("Prosa & Pesca - CStock")
        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x250+500+150")
        self.root.resizable(False, False)
        self.root.configure(bg="white")
        self.root.iconbitmap("images/system_logo.ico")

        # Frames da janela
        self.first_screen = Frame(self.root, bg="white")
        self.stock_screen = Frame(self.root, bg="white")
        self.stock_register_screen = Frame(self.root, bg="white")
        self.stock_consult_screen = Frame(self.root, bg="white")
        self.purchase_screen = Frame(self.root, bg="white")
        self.balance_screen = Frame(self.root, bg="white")

        # Botões das Principais Funções
        if True:
            self.register = PhotoImage(file=r"images\registrar_compra.png")
            self.access = PhotoImage(file=r"images\acessar_estoque.png")
            self.balance = PhotoImage(file=r"images\balancete.png")

            self.register_purchase = Button(self.first_screen, image=self.register, borderwidth=0.5, bg="white",
                                            activebackground="white", cursor="hand2", command=self.open_purchase)
            self.access_stock = Button(self.first_screen, image=self.access, borderwidth=0.5, bg="white",
                                       activebackground="white", cursor="hand2", command=self.open_stock)
            self.balance_sheet = Button(self.first_screen, image=self.balance, borderwidth=0.5, bg="white",
                                        activebackground="white", cursor="hand2", command=self.open_balance_sheet)

        # Botões do Estoque
        if True:
            self.consult = PhotoImage(file=r"images\consultar_estoque.png")
            self.register_product = PhotoImage(file=r"images\cadastrar_produto.png")

            self.access_stock_consult = Button(self.stock_screen, image=self.consult, borderwidth=0.5, bg="white",
                                               activebackground="white", cursor="hand2",
                                               command=self.open_consult_products)
            self.access_stock_register_product = Button(self.stock_screen, image=self.register_product, borderwidth=0.5,
                                                        bg="white", activebackground="white", cursor="hand2",
                                                        command=self.open_register_product)
            self.leave_stock_area = Button(self.stock_screen, text="<-", borderwidth=0.5, bg="white",
                                           activebackground="white", cursor="hand2", command=self.close_stock)

        # Botões/Campos do cadastro do produto
        if True:
            self.text_register_id = Label(self.stock_register_screen, text="ID Produto", bg="white",
                                          font=("Courier", 34))
            self.entry_register_id = Entry(self.stock_register_screen, font=("Courier", 28), bg="#DCDCDC")
            self.text_register_nome = Label(self.stock_register_screen, text="Nome Produto", bg="white",
                                            font=("Courier", 34))
            self.entry_register_nome = Entry(self.stock_register_screen, font=("Courier", 28), bg="#DCDCDC")
            self.text_register_qtd = Label(self.stock_register_screen, text="Qtd. Estoque", bg="white",
                                           font=("Courier", 34))
            self.entry_register_qtd = Entry(self.stock_register_screen, font=("Courier", 28), bg="#DCDCDC")
            self.text_register_valor_pago = Label(self.stock_register_screen, text="Valor Pago", bg="white",
                                                  font=("Courier", 34))
            self.entry_register_valor_pago = Entry(self.stock_register_screen, font=("Courier", 28), bg="#DCDCDC")
            self.text_register_valor_venda = Label(self.stock_register_screen, text="Valor Venda Un.", bg="white",
                                                   font=("Courier", 34))
            self.entry_register_valor_venda = Entry(self.stock_register_screen, font=("Courier", 28), bg="#DCDCDC")
            self.leave_register_product_area = Button(self.stock_register_screen, text="<-", borderwidth=0.5,
                                                      bg="white", activebackground="white", cursor="hand2",
                                                      command=self.close_register_product)
            self.confirm_register_product = Button(self.stock_register_screen, text="Registrar", borderwidth=0.5,
                                                   bg="white", activebackground="white", cursor="hand2",
                                                   font=("Courier", 20), command=self.register_product_on_stock)

        # Botões/Dados da consulta de produtos
        if True:
            self.canvas_consult_area = Canvas(self.stock_consult_screen,
                                              width=self.stock_consult_screen.winfo_width(),
                                              height=self.stock_consult_screen.winfo_height(),
                                              bg="white")
            self.scroll_consult_product_area = Scrollbar(self.stock_consult_screen, orient=VERTICAL)
            self.scroll_consult_product_area.pack(side=RIGHT, fill=Y)
            self.scroll_consult_product_area.config(command=self.canvas_consult_area.yview)

            self.canvas_consult_area.config(yscrollcommand=self.scroll_consult_product_area.set)

            self.leave_consult_product_area = Button(self.canvas_consult_area, text="<-", borderwidth=0.5,
                                                     bg="white", activebackground="white", cursor="hand2",
                                                     command=self.close_consult_products)
            self.text_consult_id = Label(self.canvas_consult_area,
                                         text="ID", bg="white", font=("Courier", 18))
            self.text_consult_nome = Label(self.canvas_consult_area,
                                           text="Nome", bg="white", font=("Courier", 18))
            self.text_consult_estoques = Label(self.canvas_consult_area,
                                               text="Estoque\n1 2 3", bg="white", font=("Courier", 18))
            self.text_consult_custos = Label(self.canvas_consult_area,
                                             text="Custo\n1 2 3", bg="white", font=("Courier", 18))
            self.text_consult_custo_total = Label(self.canvas_consult_area,
                                                  text="Custo\nTotal", bg="white", font=("Courier", 18))
            self.text_consult_valor_venda = Label(self.canvas_consult_area,
                                                  text="Valor\nVenda Un.", bg="white", font=("Courier", 18))
            self.text_consult_lucro = Label(self.canvas_consult_area,
                                            text="Lucro\nPotencial", bg="white", font=("Courier", 18))
            self.sep = ttk.Separator(orient="horizontal")
            self.sep.place(in_=self.text_consult_id, x=0, relx=-3.5, rely=1.7, height=2, relwidth=38.17)

        # Botões/Campos/Dados do registro de compra
        if True:
            self.leave_register_purchase_area = Button(self.purchase_screen, text="<-", borderwidth=0.5,
                                                       bg="white", activebackground="white", cursor="hand2",
                                                       command=self.close_purchase)
            self.text_purchase_id = Label(self.purchase_screen, text="ID Produto", bg="white",
                                          font=("Courier", 25))
            self.entry_purchase_id = Entry(self.purchase_screen, font=("Courier", 28), bg="#DCDCDC")

            self.check_id = PhotoImage(file=r"images\check_button.png")
            self.check_purchase_id = Button(self.purchase_screen, image=self.check_id, borderwidth=0.5, bg="white",
                                            cursor="hand2", activebackground="white", command=self.check_purchase_id)
            self.text_purchase_nome = Label(self.purchase_screen, text="Nome Produto", bg="white",
                                            font=("Courier", 25))
            self.entry_purchase_nome = Entry(self.purchase_screen, font=("Courier", 28), bg="#DCDCDC", state="disabled")

            self.area_purchase_stock = Frame(self.purchase_screen, bg="white")
            self.text_purchase_stock = Label(self.area_purchase_stock, text="Qtd. Estoque", bg="white",
                                             font=("Courier", 25))
            self.entry_purchase_stock = Entry(self.area_purchase_stock, font=("Courier", 25), bg="#DCDCDC", width=12,
                                              state="disabled")
            self.text_purchase_price = Label(self.area_purchase_stock, text="x R$", bg="white",
                                             font=("Courier", 25))
            self.entry_purchase_price = Entry(self.area_purchase_stock, font=("Courier", 25), bg="#DCDCDC", width=6)
            self.text_purchase_quantity = Label(self.area_purchase_stock, text="Quantidade", bg="white",
                                                font=("Courier", 25))
            self.entry_purchase_quantity = Entry(self.area_purchase_stock, font=("Courier", 25), bg="#DCDCDC", width=10)
            self.btt_register_purchase = Button(self.purchase_screen, text="Registrar", font=("Courier", 25),
                                                borderwidth=0.5, bg="white", cursor="hand2",
                                                activebackground="white", command=self.confirm_purchase)

        # Botões de consultar balancete
        if True:
            self.canvas_balance_area = Canvas(self.balance_screen, width=self.balance_screen.winfo_width(),
                                              height=self.balance_screen.winfo_height(), bg="white")
            self.scroll_balance_area = Scrollbar(self.balance_screen, orient=VERTICAL)
            self.scroll_balance_area.pack(side=RIGHT, fill=Y)
            self.scroll_balance_area.config(command=self.canvas_balance_area.yview)
            self.canvas_balance_area.config(yscrollcommand=self.scroll_balance_area.set)
            self.leave_balance_area = Button(self.canvas_balance_area, text="<-", borderwidth=0.5,
                                             bg="white", activebackground="white", cursor="hand2",
                                             command=self.close_balance_sheet)
            self.text_balance_nome = Label(self.canvas_balance_area, text="Nome", bg="white", font=("Courier", 18))
            self.text_balance_custo = Label(self.canvas_balance_area, text="Custo", bg="white", font=("Courier", 18))
            self.text_balance_lucro = Label(self.canvas_balance_area, text="Renda\nBruta", bg="white", font=("Courier", 18))
            ttk.Separator(orient="horizontal").place(
                in_=self.text_balance_nome, x=0, relx=-1.56, rely=1.7, height=2, relwidth=10.85)

        self.register_purchase.grid(row=1, column=1, ipadx=5, padx=40, pady=30)
        self.access_stock.grid(row=1, column=2, ipadx=5, padx=40, pady=30)
        self.balance_sheet.grid(row=1, column=3, ipadx=5, padx=40, pady=30)

        self.access_stock_consult.grid(row=1, column=1, ipadx=5, padx=77, pady=0)
        self.access_stock_register_product.grid(row=1, column=2, ipadx=5, padx=80, pady=0)
        self.leave_stock_area.grid(row=0, column=0, padx=5, pady=5)

        self.leave_register_product_area.grid(row=0, column=0, padx=(5, 65), pady=5)
        self.text_register_id.grid(row=1, column=1, padx=5, pady=3, sticky=W)
        self.entry_register_id.grid(row=2, column=1, padx=5, pady=5, sticky=W)
        self.text_register_nome.grid(row=3, column=1, padx=5, pady=3, sticky=W)
        self.entry_register_nome.grid(row=4, column=1, padx=5, pady=5, sticky=W)
        self.text_register_qtd.grid(row=5, column=1, padx=5, pady=3, sticky=W)
        self.entry_register_qtd.grid(row=6, column=1, padx=5, pady=5, sticky=W)
        self.text_register_valor_pago.grid(row=7, column=1, padx=5, pady=3, sticky=W)
        self.entry_register_valor_pago.grid(row=8, column=1, padx=5, pady=5, sticky=W)
        self.text_register_valor_venda.grid(row=9, column=1, padx=5, pady=3, sticky=W)
        self.entry_register_valor_venda.grid(row=10, column=1, padx=5, pady=5, sticky=W)
        self.confirm_register_product.grid(row=11, column=1, pady=15)

        self.canvas_consult_area.pack(side=LEFT, expand=True, fill=BOTH)
        self.leave_consult_product_area.grid(row=0, column=0, padx=5, pady=5, sticky=W)
        self.canvas_consult_area.create_window(110, 40, window=self.text_consult_id)
        self.canvas_consult_area.create_window(380, 40, window=self.text_consult_nome)
        self.canvas_consult_area.create_window(560, 40, window=self.text_consult_estoques)
        self.canvas_consult_area.create_window(680, 40, window=self.text_consult_custos)
        self.canvas_consult_area.create_window(790, 40, window=self.text_consult_custo_total)
        self.canvas_consult_area.create_window(900, 40, window=self.text_consult_valor_venda)
        self.canvas_consult_area.create_window(1040, 40, window=self.text_consult_lucro)

        self.leave_register_purchase_area.grid(row=0, column=0, padx=(5, 0), pady=5, sticky=W)
        self.text_purchase_id.grid(row=1, column=1, padx=(65, 0), pady=3, sticky=W)
        self.entry_purchase_id.grid(row=2, column=1, padx=(65, 0), pady=5, sticky=W)
        self.check_purchase_id.grid(row=2, column=2, padx=5, pady=5, sticky=W)
        self.text_purchase_nome.grid(row=3, column=1, padx=(65, 0), pady=(10, 5), sticky=W)
        self.entry_purchase_nome.grid(row=4, column=1, padx=(65, 0), sticky=W)
        self.area_purchase_stock.grid(row=5, column=1, padx=(65, 0), pady=(10, 0), sticky=W)
        self.text_purchase_stock.grid(row=0, column=0, sticky=W)
        self.entry_purchase_stock.grid(row=1, column=0, sticky=W)
        self.text_purchase_quantity.grid(row=2, column=0, pady=(10, 0), sticky=W)
        self.entry_purchase_quantity.grid(row=3, column=0, sticky=W)
        self.text_purchase_price.grid(row=3, column=0, padx=(210, 0), sticky=W)
        self.entry_purchase_price.grid(row=3, column=0, padx=(300, 0))
        self.btt_register_purchase.grid(row=6, column=1, padx=(80, 0), pady=(25, 0))

        self.canvas_balance_area.pack(side=LEFT, expand=True, fill=BOTH)
        self.leave_balance_area.grid(row=0, column=0, padx=(5, 0), pady=5, sticky=W)
        self.canvas_balance_area.create_window(120, 40, window=self.text_balance_nome)
        self.canvas_balance_area.create_window(300, 40, window=self.text_balance_custo)
        self.canvas_balance_area.create_window(500, 40, window=self.text_balance_lucro)

        self.first_screen.pack()

    def open_purchase(self):
        # Reconfiguração da janela inicial para modelo compra
        self.first_screen.pack_forget()
        self.purchase_screen.pack(side=TOP, anchor=NW)
        self.root.title("Prosa & Pesca - Compra - CStock")
        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x550+500+100")

    def close_purchase(self):
        # Reconfiguração da janela compra para modelo inicial
        self.first_screen.pack()
        self.purchase_screen.pack_forget()

        self.entry_purchase_id.delete(0, END)
        self.entry_purchase_nome.delete(0, END)
        self.entry_purchase_stock.delete(0, END)
        self.entry_purchase_price.delete(0, END)
        self.entry_purchase_quantity.delete(0, END)

        self.entry_purchase_id.configure(bg="#DCDCDC")
        self.entry_purchase_nome.configure(bg="#DCDCDC")
        self.entry_purchase_stock.configure(bg="#DCDCDC")
        self.entry_purchase_price.configure(bg="#DCDCDC")
        self.entry_purchase_quantity.configure(bg="#DCDCDC")

        self.root.title("Prosa & Pesca - CStock")
        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x250+500+150")

    def check_purchase_id(self):
        df = pd.DataFrame(pd.read_excel(CAMINHO, sheet_name="Estoque"),
                          columns=["ID produto (Código de Barras)", "Nome", "Quantidade em Estoque 1",
                                   "Quantidade em Estoque 2", "Quantidade em Estoque 3",
                                   "Valor de Venda"]).values.tolist()
        if len(self.entry_purchase_id.get()) < 1:
            self.entry_purchase_id.configure(bg="#ff4c4c")
            return False
        else:
            for index in range(len(df)):
                if df[index][0] == int(self.entry_purchase_id.get()):
                    self.entry_purchase_id.configure(bg="#93bf85")
                    self.entry_purchase_nome.configure(state="normal")
                    self.entry_purchase_nome.delete(0, END)
                    self.entry_purchase_nome.insert(0, df[index][1])
                    self.entry_purchase_nome.configure(state="disabled")

                    self.entry_purchase_stock.configure(state="normal")
                    self.entry_purchase_stock.delete(0, END)
                    self.entry_purchase_stock.insert(0, df[index][2] + df[index][3] + df[index][4])
                    self.entry_purchase_stock.configure(state="disabled")

                    self.entry_purchase_price.delete(0, END)
                    self.entry_purchase_price.insert(0, df[index][5])

    def confirm_purchase(self):
        if int(self.entry_purchase_stock.get()) >= int(self.entry_purchase_quantity.get()) > 0:
            self.entry_purchase_quantity.configure(bg="#DCDCDC")

            df = pd.DataFrame(pd.read_excel(CAMINHO, sheet_name="Estoque"),
                              columns=["ID produto (Código de Barras)"]).values.tolist()

            wb = load_workbook(CAMINHO)
            ws = wb["Estoque"]
            row_index = 0

            for row in range(len(df)):
                if df[row][0] == int(self.entry_purchase_id.get()):
                    row_index = row + 2

            ws.cell(row=row_index, column=11).value = \
                int(ws.cell(row=row_index, column=11).value) + int(self.entry_purchase_quantity.get())
            ws.cell(row=row_index, column=12).value = float(ws.cell(row=row_index, column=12).value) + \
                float(self.entry_purchase_price.get()) * int(self.entry_purchase_quantity.get())

            if int(ws.cell(row=row_index, column=3).value) >= int(self.entry_purchase_quantity.get()):
                ws.cell(row=row_index, column=3).value = \
                    int(ws.cell(row=row_index, column=3).value) - int(self.entry_purchase_quantity.get())
            elif int(ws.cell(row=row_index, column=4).value) >= int(self.entry_purchase_quantity.get()):
                ws.cell(row=row_index, column=4).value = \
                    int(ws.cell(row=row_index, column=4).value) - int(self.entry_purchase_quantity.get())
            elif int(ws.cell(row=row_index, column=5).value) >= int(self.entry_purchase_quantity.get()):
                ws.cell(row=row_index, column=5).value = \
                    int(ws.cell(row=row_index, column=5).value) - int(self.entry_purchase_quantity.get())
            else:
                # Retirar valor comprado de multiplos estoques
                ws.cell(row=row_index, column=3).value = 0
                negative_one = -(int(ws.cell(row=row_index, column=3).value) - int(self.entry_purchase_quantity.get()))
                if (int(ws.cell(row=row_index, column=4).value) - negative_one) >= 0:
                    ws.cell(row=row_index, column=4).value = int(ws.cell(row=row_index, column=4).value) - negative_one
                else:
                    ws.cell(row=row_index, column=4).value = 0
                    negative_two = \
                        -(int(ws.cell(row=row_index, column=4).value) - int(self.entry_purchase_quantity.get()))
                    if (int(ws.cell(row=row_index, column=5).value) - negative_two) >= 0:
                        ws.cell(row=row_index, column=5).value = \
                            int(ws.cell(row=row_index, column=5).value) - negative_two
                # Reset custo unitario em caso de estoque vazio
                if ws.cell(row=row_index, column=3).value == 0:
                    ws.cell(row=row_index, column=6).value = 0.0
                if ws.cell(row=row_index, column=4).value == 0:
                    ws.cell(row=row_index, column=7).value = 0.0
                if ws.cell(row=row_index, column=5).value == 0:
                    ws.cell(row=row_index, column=8).value = 0.0
            wb.save(CAMINHO)

            # Reset campos
            self.entry_purchase_id.delete(0, END)
            self.entry_purchase_nome.delete(0, END)
            self.entry_purchase_stock.delete(0, END)
            self.entry_purchase_price.delete(0, END)
            self.entry_purchase_quantity.delete(0, END)

            self.entry_purchase_id.configure(bg="#DCDCDC")
            self.entry_purchase_nome.configure(bg="#DCDCDC")
            self.entry_purchase_stock.configure(bg="#DCDCDC")
            self.entry_purchase_price.configure(bg="#DCDCDC")
            self.entry_purchase_quantity.configure(bg="#DCDCDC")
        else:
            self.entry_purchase_quantity.configure(bg="#ff4c4c")

    def open_stock(self):
        # Reconfiguração da janela inicial para modelo estoque
        self.first_screen.pack_forget()
        self.stock_screen.pack(side=TOP, anchor=NW)
        self.root.title("Prosa & Pesca - Estoque - CStock")

    def close_stock(self):
        # Reconfiguração da janela estoque para modelo inicial
        self.first_screen.pack()
        self.stock_screen.pack_forget()
        self.root.title("Prosa & Pesca - CStock")

    def open_consult_products(self):
        # Reconfiguração da janela estoque para modelo consulta
        self.stock_screen.pack_forget()
        self.stock_consult_screen.pack(expand=True, fill=BOTH)
        self.get_stock_products()
        self.root.geometry(str(int(self.root.winfo_screenwidth() / 1.35)) + "x680+200+50")

    def close_consult_products(self):
        # Reconfiguração da janela consulta para modelo estoque
        self.stock_screen.pack(side=TOP, anchor=NW)
        self.stock_consult_screen.pack_forget()

        for index in range(8, len(self.canvas_consult_area.winfo_children()), 1):
            self.canvas_consult_area.winfo_children()[-1].destroy()
            self.canvas_consult_area.winfo_children().pop(-1)

        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x250+500+150")

    def get_stock_products(self):
        df = pd.DataFrame(pd.read_excel(CAMINHO, sheet_name="Estoque"),
                          columns=["ID produto (Código de Barras)", "Nome", "Quantidade em Estoque 1",
                                   "Quantidade em Estoque 2", "Quantidade em Estoque 3", "Valor Pago 1",
                                   "Valor Pago 2", "Valor Pago 3", "Valor de Venda", "Custo Total"]).values.tolist()
        x1 = 95
        for row in range(len(df)):
            label_for_separator = Label(self.canvas_consult_area, bg="white", text=df[row][0], font=("Courier", 18))
            ttk.Separator(orient="horizontal").place(in_=label_for_separator,
                                                     x=0, relx=-3.5, rely=2.7, height=2, relwidth=9.57)
            self.canvas_consult_area.create_window(110, x1 + 40, window=label_for_separator)
            self.canvas_consult_area.create_window(380, x1 + 40,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][1], font=("Courier", 18)))
            self.canvas_consult_area.create_window(560, x1,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][2], font=("Courier", 18)))
            self.canvas_consult_area.create_window(560, x1 + 40,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][3], font=("Courier", 18)))
            self.canvas_consult_area.create_window(560, x1 + 80,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][4], font=("Courier", 18)))
            self.canvas_consult_area.create_window(680, x1,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][5], font=("Courier", 18)))
            self.canvas_consult_area.create_window(680, x1 + 40,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][6], font=("Courier", 18)))
            self.canvas_consult_area.create_window(680, x1 + 80,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][7], font=("Courier", 18)))
            self.canvas_consult_area.create_window(790, x1 + 40,
                                                   window=Label(self.canvas_consult_area, bg="white",
                                                                text=df[row][9], font=("Courier", 18)))
            self.canvas_consult_area.create_window(900, x1 + 40,
                                                   window=Label(self.canvas_consult_area,
                                                                bg="white", text=df[row][8], font=("Courier", 18)))
            self.canvas_consult_area.create_window(1040, x1 + 40,
                                                   window=Label(self.canvas_consult_area, bg="white",
                                                                text=(df[row][2] * df[row][8] +
                                                                      df[row][3] * df[row][8] +
                                                                      df[row][4] * df[row][8]) -
                                                                     (df[row][7] + df[row][6] + df[row][5]),
                                                                font=("Courier", 18)))
            x1 += 130
        self.canvas_consult_area.configure(scrollregion=self.canvas_consult_area.bbox("all"))

    def open_register_product(self):
        # Reconfiguração da janela estoque para modelo cadastro
        self.stock_screen.pack_forget()
        self.stock_register_screen.pack(side=TOP, anchor=NW)
        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x680+500+50")

    def close_register_product(self):
        # Reconfiguração da janela cadastro para modelo estoque
        self.stock_screen.pack(side=TOP, anchor=NW)
        self.stock_register_screen.pack_forget()

        # Limpar campo
        self.entry_register_id.delete(0, END)
        self.entry_register_nome.delete(0, END)
        self.entry_register_qtd.delete(0, END)
        self.entry_register_valor_venda.delete(0, END)
        self.entry_register_valor_pago.delete(0, END)

        # Resetar cor do campo
        self.entry_register_id.configure(bg="#DCDCDC")
        self.entry_register_nome.configure(bg="#DCDCDC")
        self.entry_register_qtd.configure(bg="#DCDCDC")
        self.entry_register_valor_venda.configure(bg="#DCDCDC")
        self.entry_register_valor_pago.configure(bg="#DCDCDC")

        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x250+500+150")

    def register_product_on_stock(self):
        if self.validate_entrys():
            df = pd.DataFrame(DATA, columns=["ID produto (Código de Barras)"]).values.tolist()

            new_product = True
            wb = load_workbook(CAMINHO)
            ws = wb["Estoque"]
            row_index = len(ws['A']) + 1

            for row in range(len(df)):
                if df[row][0] == int(self.entry_register_id.get()):
                    new_product = False
                    row_index = row + 2

            if new_product:
                ws.cell(row=row_index, column=1).value = int(self.entry_register_id.get())
                ws.cell(row=row_index, column=2).value = self.entry_register_nome.get()
                ws.cell(row=row_index, column=3).value = int(self.entry_register_qtd.get())
                ws.cell(row=row_index, column=4).value = 0
                ws.cell(row=row_index, column=5).value = 0
                ws.cell(row=row_index, column=6).value = float(self.entry_register_valor_pago.get())
                ws.cell(row=row_index, column=7).value = 0.0
                ws.cell(row=row_index, column=8).value = 0.0
                ws.cell(row=row_index, column=9).value = float(self.entry_register_valor_pago.get())
                ws.cell(row=row_index, column=10).value = float(self.entry_register_valor_venda.get())
            else:
                if ws.cell(row=row_index, column=6).value == float(self.entry_register_valor_pago.get()):
                    ws.cell(row=row_index, column=3).value = \
                        int(ws.cell(row=row_index, column=3).value + int(self.entry_register_qtd.get()))
                elif ws.cell(row=row_index, column=7).value == float(self.entry_register_valor_pago.get()):
                    ws.cell(row=row_index, column=4).value = \
                        int(ws.cell(row=row_index, column=4).value + int(self.entry_register_qtd.get()))
                elif ws.cell(row=row_index, column=8).value == float(self.entry_register_valor_pago.get()):
                    ws.cell(row=row_index, column=5).value = \
                        int(ws.cell(row=row_index, column=5).value + int(self.entry_register_qtd.get()))
                else:
                    if ws.cell(row=row_index, column=3).value == 0:
                        ws.cell(row=row_index, column=3).value = int(self.entry_register_qtd.get())
                        ws.cell(row=row_index, column=6).value = float(self.entry_register_valor_pago.get())
                    elif ws.cell(row=row_index, column=4).value == 0:
                        ws.cell(row=row_index, column=4).value = int(self.entry_register_qtd.get())
                        ws.cell(row=row_index, column=7).value = float(self.entry_register_valor_pago.get())
                    elif ws.cell(row=row_index, column=5).value == 0:
                        ws.cell(row=row_index, column=5).value = int(self.entry_register_qtd.get())
                        ws.cell(row=row_index, column=8).value = float(self.entry_register_valor_pago.get())
                ws.cell(row=row_index, column=9).value = \
                    ws.cell(row=row_index, column=9).value + float(self.entry_register_valor_pago.get())
                ws.cell(row=row_index, column=10).value = float(self.entry_register_valor_venda.get())
            wb.save(CAMINHO)

    def validate_entrys(self):
        if not self.entry_register_id.get().isnumeric():
            self.entry_register_id.configure(bg="#ff4c4c")
            return False
        else:
            self.entry_register_id.configure(bg="#93bf85")

        df = pd.DataFrame(DATA, columns=["ID produto (Código de Barras)", "Nome"]).values.tolist()
        for row in range(len(df)):
            if df[row][0] == int(self.entry_register_id.get()):
                if len(self.entry_register_nome.get()) > 1:
                    if df[row][1] == self.entry_register_nome.get():
                        self.entry_register_nome.configure(bg="#93bf85")
                    else:
                        self.entry_register_nome.delete(0, END)
                        self.entry_register_nome.insert(0, df[row][1])
                        self.entry_register_nome.configure(bg="#ffff56")
                        return False
                else:
                    self.entry_register_nome.configure(bg="#93bf85")
                    return False
        if not self.entry_register_qtd.get().isnumeric():
            self.entry_register_qtd.configure(bg="#ff4c4c")
            return False
        else:
            self.entry_register_qtd.configure(bg="#93bf85")
        if not self.is_number(self.entry_register_valor_pago.get()):
            self.entry_register_valor_pago.configure(bg="#ff4c4c")
            return False
        else:
            self.entry_register_valor_pago.configure(bg="#93bf85")
        if not self.is_number(self.entry_register_valor_venda.get()):
            self.entry_register_valor_venda.configure(bg="#ff4c4c")
            return False
        else:
            self.entry_register_valor_venda.configure(bg="#93bf85")
        return True

    def open_balance_sheet(self):
        # Reconfiguração da janela inicial para modelo balancete
        self.first_screen.pack_forget()
        self.balance_screen.pack(expand=True, fill=BOTH)
        self.root.title("Prosa & Pesca - Balancete - CStock")
        self.get_balance_data()
        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x550+500+100")

    def close_balance_sheet(self):
        # Reconfiguração da janela balancete para modelo inicial
        self.first_screen.pack()
        self.balance_screen.pack_forget()
        self.root.title("Prosa & Pesca - CStock")
        self.root.geometry(str(int(self.root.winfo_screenwidth() / 2.38)) + "x250+500+150")

    def get_balance_data(self):
        df = pd.DataFrame(pd.read_excel(CAMINHO, sheet_name="Estoque"),
                          columns=["Nome", "Custo Total", "Rendimento Bruto"]).values.tolist()

        x1 = 120
        for row in range(len(df)):
            label_for_separator = Label(self.canvas_balance_area, bg="white", text=df[row][0], font=("Courier", 18))
            ttk.Separator(orient="horizontal").place(in_=label_for_separator,
                                                     x=0, relx=-5.0, rely=2.0, height=2, relwidth=30.0)
            self.canvas_balance_area.create_window(120, x1, window=label_for_separator)

            self.canvas_balance_area.create_window(300, x1, window=Label(
                self.canvas_balance_area, bg="white", text=df[row][1], font=("Courier", 18)))
            self.canvas_balance_area.create_window(500, x1, window=Label(
                self.canvas_balance_area, bg="white", text=df[row][2], font=("Courier", 18)))
            x1 += 80
        self.canvas_balance_area.create_window(500, x1, window=Label(
            self.canvas_balance_area, bg="white", text="0", font=("Courier", 18)))
        self.canvas_balance_area.configure(scrollregion=self.canvas_balance_area.bbox("all"))

    @staticmethod
    def is_number(string):
        try:
            float(string)
            return True
        except ValueError:
            return False


orq = ManagerInterface()
orq.root.mainloop()
